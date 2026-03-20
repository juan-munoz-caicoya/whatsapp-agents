"""
Crea un archivo Excel (.xlsx) con los datos fake para la demo P2.
Google Sheets importa .xlsx directamente manteniendo las hojas.

Uso:
  python create_sheets.py

Output:
  autoescuela_demo.xlsx (en la misma carpeta)

Después subilo a Google Drive y abrilo con Google Sheets.
"""

import subprocess
import sys

# Instalar openpyxl si no está
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# --- Estilos ---
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def style_data(ws, num_rows, num_cols):
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4


# ============================================================
# HOJA 1: agenda
# ============================================================
ws_agenda = wb.active
ws_agenda.title = "agenda"

agenda_headers = [
    "alumno_nombre",
    "alumno_telefono",
    "fecha",
    "hora",
    "instructor",
    "punto_recogida",
    "recordatorio_enviado",
    "estado",
]

agenda_data = [
    ["María López", "+34600111222", "2026-03-21", "10:00", "Antonio", "Av. Andalucía 15", "no", "pendiente"],
    ["Carlos Ruiz", "+34600333444", "2026-03-21", "11:00", "Antonio", "Plaza Mayor 3", "no", "pendiente"],
    ["Laura García", "+34600555666", "2026-03-21", "12:00", "Marta", "Av. Andalucía 15", "no", "pendiente"],
    ["Pedro Sánchez", "+34600777888", "2026-03-22", "09:00", "Antonio", "C/ Larios 8", "no", "pendiente"],
]

for col_idx, header in enumerate(agenda_headers, 1):
    ws_agenda.cell(row=1, column=col_idx, value=header)

for row_idx, row_data in enumerate(agenda_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_agenda.cell(row=row_idx, column=col_idx, value=value)

style_header(ws_agenda, len(agenda_headers))
style_data(ws_agenda, len(agenda_data) + 1, len(agenda_headers))
auto_width(ws_agenda)

# ============================================================
# HOJA 2: disponibilidad
# ============================================================
ws_disp = wb.create_sheet("disponibilidad")

disp_headers = [
    "fecha",
    "hora",
    "instructor",
    "punto_recogida",
    "ocupado",
]

disp_data = [
    ["2026-03-22", "10:00", "Antonio", "Av. Andalucía 15", "no"],
    ["2026-03-22", "12:00", "Marta", "Plaza Mayor 3", "no"],
    ["2026-03-23", "09:00", "Antonio", "Av. Andalucía 15", "no"],
    ["2026-03-23", "11:00", "Antonio", "Plaza Mayor 3", "no"],
    ["2026-03-24", "10:00", "Marta", "C/ Larios 8", "no"],
    ["2026-03-24", "12:00", "Antonio", "Av. Andalucía 15", "no"],
]

for col_idx, header in enumerate(disp_headers, 1):
    ws_disp.cell(row=1, column=col_idx, value=header)

for row_idx, row_data in enumerate(disp_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_disp.cell(row=row_idx, column=col_idx, value=value)

style_header(ws_disp, len(disp_headers))
style_data(ws_disp, len(disp_data) + 1, len(disp_headers))
auto_width(ws_disp)

# ============================================================
# HOJA 3: log
# ============================================================
ws_log = wb.create_sheet("log")

log_headers = [
    "timestamp",
    "alumno",
    "accion",
    "clase_original",
    "nueva_clase",
]

for col_idx, header in enumerate(log_headers, 1):
    ws_log.cell(row=1, column=col_idx, value=header)

style_header(ws_log, len(log_headers))
auto_width(ws_log)

# ============================================================
# Guardar
# ============================================================
output_path = Path(__file__).parent / "autoescuela_demo.xlsx"
wb.save(output_path)
print(f"Archivo creado: {output_path}")
print("Subilo a Google Drive y abrilo con Google Sheets.")
